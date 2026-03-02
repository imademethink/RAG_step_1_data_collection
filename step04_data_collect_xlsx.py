import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference

# tell top 10 operations to be done on Excel file using Python

# 1. Reading Excel Files
# 2. Writing to Excel
# 3. Managing Multiple Sheets
# 4. Cell-Level Editing
# 5. Formatting and Styling
# 6. Filtering and Sorting
# 7. Aggregating Data
# 8. Creating Pivot Tables
# 9. Inserting Formulas
# 10. Adding Charts and Images




# 1. Reading Excel Files
df1 = pd.read_excel('data.xlsx', sheet_name='Sheet1')


# 2. Writing to Excel
df2 = pd.read_excel('data.xlsx', sheet_name='Sheet1')
df2.to_excel('output.xlsx', index=False)


# 3. Managing Multiple Sheets
wb123 = load_workbook('data.xlsx')
print(wb123.sheetnames)  # Output: ['Sheet1', 'Sheet2', 'Sales']
# create a sheet
ws_new = wb123.create_sheet(title="Marketing", index=0)
# access specific worksheet
ws_sales = wb123["Sales"]
# rename worksheet
ws_new.title = "January_Data"
# delete worksheet
wb123.remove(wb123["Sheet2"])


# 4. Cell-Level Editing
wb99 = load_workbook('data.xlsx')
ws99 = wb99.active
ws99['A1'] = "New Value"






# 5. Formatting and Styling
# 1. Initialize
wb = Workbook()
ws = wb.active
ws.title = "StyledSheet"

# 2. Add Data
header = ["Project Name", "Budget", "Status"]
ws.append(header)
ws.append(["AI Engine", 50000, "In Progress"])

# 3. Define Styles
# Blue background fill
header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
# White, bold font
header_font = Font(color="FFFFFF", bold=True, size=12)
# Center alignment
center_align = Alignment(horizontal="center", vertical="center")
# Thin black border
thin_side = Side(style="thin", color="000000")
square_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
# 4. Apply Styles to the header row (Row 1)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = square_border
# 5. Save the file
wb.save("StyledReport.xlsx")






# 6. Filtering and Sorting
# Load the Excel file
df = pd.read_excel('data.xlsx')
# FILTERING: Select rows where 'Sales' > 1000 AND 'Status' is 'Active'
filtered_df = df[(df['Sales'] > 1000) & (df['Status'] == 'Active')]
# SORTING: Sort by 'Region' (ascending) and then 'Sales' (descending)
sorted_df = filtered_df.sort_values(by=['Region', 'Sales'], ascending=[True, False])
# Save the result to a new Excel file
sorted_df.to_excel('filtered_and_sorted.xlsx', index=False)








# 7. Aggregating Data
# 1. Load the Excel file
df = pd.read_excel('company_data.xlsx')
# 2. Basic Aggregation: Calculate stats for the whole column
total_budget = df['Salary'].sum()
average_salary = df['Salary'].mean()
employee_count = df['Employee'].count()
# 3. Grouped Aggregation: Stats per Department
# We group by 'Department' and apply multiple functions to 'Salary'
dept_summary = df.groupby('Department')['Salary'].agg(['sum', 'mean', 'count'])
# 4. Rename columns for clarity
dept_summary.columns = ['Total Salary', 'Average Salary', 'Headcount']
# 5. Save the summary to a new sheet
dept_summary.to_excel('department_summary.xlsx')
print(dept_summary)




# 8. Creating Pivot Tables
# 1. Load your Excel data
df = pd.read_excel('sales_data.xlsx')
# 2. Create the Pivot Table
# values: the column to aggregate (Revenue)
# index: the rows of the pivot table (Region)
# columns: the horizontal headers (Product)
# aggfunc: the calculation to perform (sum, mean, count, etc.)
pivot = pd.pivot_table(df,
                       values='Revenue',
                       index='Region',
                       columns='Product',
                       aggfunc='sum',
                       fill_value=0) # Replaces empty cells with 0
# 3. Add a "Grand Total" column and row (Optional)
pivot.loc['Grand Total'] = pivot.sum()
pivot['Total Revenue'] = pivot.sum(axis=1)
# 4. Save to a new Excel file
pivot.to_excel('sales_pivot_report.xlsx')
print(pivot)




# 9. Inserting Formulas
wb = Workbook()
ws = wb.active
# Add some sample data
ws['A1'] = 100
ws['A2'] = 200
# Insert a SUM formula
ws['A3'] = "=SUM(A1:A2)"
# Save the workbook
wb.save("formulas.xlsx")




# 10. Adding Charts and Images
wb = Workbook()
ws = wb.active
# Load your image file
img = Image('logo.png')
# Anchor the image to a specific cell
img.anchor = 'B2'
# Add image to the worksheet
ws.add_image(img)
wb.save('ImageReport.xlsx')







wb = Workbook()
ws = wb.active
# 1. Add sample data
data = [
    ["Category", "Sales"],
    ["Electronics", 4500],
    ["Home", 3200],
    ["Fashion", 2100],
]
for row in data:
    ws.append(row)
# 2. Create the chart
chart = BarChart()
chart.title = "Sales by Category"
chart.y_axis.title = 'Revenue ($)'
chart.x_axis.title = 'Category'
# 3. Define the data range (Sales values in column 2, rows 1-4)
values = Reference(ws, min_col=2, min_row=1, max_row=4, max_col=2)
chart.add_data(values, titles_from_data=True)
# 4. Add chart to a specific cell
ws.add_chart(chart, "E2")
wb.save("ChartReport.xlsx")
